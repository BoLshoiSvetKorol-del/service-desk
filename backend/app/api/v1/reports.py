import csv
import io
from datetime import date, datetime, timezone
from typing import Literal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, case, text

from app.database import get_db
from app.models.ticket import Ticket, TicketStatus
from app.models.priority import Priority
from app.models.ticket_type import TicketType
from app.models.user import User, UserRole
from app.models.department import Department
from app.schemas.report import (
    TicketsCountItem, TicketsCountResponse,
    StatusDistributionItem, AvgResolutionItem, SLAComplianceResponse,
)
from app.utils.permissions import require_role

router = APIRouter()

_ALLOWED_ROLES = (UserRole.admin, UserRole.agent)


def _base_filters(
    date_from: date | None,
    date_to: date | None,
    department_id: int | None,
    type_id: int | None,
    priority_id: int | None,
):
    """Общие WHERE-условия для всех отчётов."""
    conditions = []
    if date_from:
        conditions.append(Ticket.created_at >= datetime(date_from.year, date_from.month, date_from.day, tzinfo=timezone.utc))
    if date_to:
        end = datetime(date_to.year, date_to.month, date_to.day, 23, 59, 59, tzinfo=timezone.utc)
        conditions.append(Ticket.created_at <= end)
    if department_id:
        conditions.append(Ticket.department_id == department_id)
    if type_id:
        conditions.append(Ticket.type_id == type_id)
    if priority_id:
        conditions.append(Ticket.priority_id == priority_id)
    return conditions


@router.get("/tickets-count", response_model=TicketsCountResponse)
async def tickets_count(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    department_id: int | None = Query(default=None),
    type_id: int | None = Query(default=None),
    priority_id: int | None = Query(default=None),
    groupby: Literal["day", "week", "month"] = Query(default="day"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(*_ALLOWED_ROLES)),
):
    period_col = func.date_trunc(groupby, Ticket.created_at).label("period")
    conditions = _base_filters(date_from, date_to, department_id, type_id, priority_id)

    query = (
        select(period_col, func.count(Ticket.id).label("count"))
        .where(and_(*conditions) if conditions else text("1=1"))
        .group_by(period_col)
        .order_by(period_col)
    )
    result = await db.execute(query)
    rows = result.all()

    items = [
        TicketsCountItem(period=str(row.period)[:10], count=row.count)
        for row in rows
    ]
    return TicketsCountResponse(items=items, total=sum(i.count for i in items))


@router.get("/by-status", response_model=list[StatusDistributionItem])
async def by_status(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    department_id: int | None = Query(default=None),
    type_id: int | None = Query(default=None),
    priority_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(*_ALLOWED_ROLES)),
):
    conditions = _base_filters(date_from, date_to, department_id, type_id, priority_id)

    query = (
        select(Ticket.status.label("status"), func.count(Ticket.id).label("count"))
        .where(and_(*conditions) if conditions else text("1=1"))
        .group_by(Ticket.status)
    )
    result = await db.execute(query)
    return [StatusDistributionItem(status=row.status.value, count=row.count) for row in result.all()]


@router.get("/avg-resolution-time", response_model=list[AvgResolutionItem])
async def avg_resolution_time(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    department_id: int | None = Query(default=None),
    type_id: int | None = Query(default=None),
    priority_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(*_ALLOWED_ROLES)),
):
    conditions = _base_filters(date_from, date_to, department_id, type_id, priority_id)
    conditions.append(Ticket.closed_at.isnot(None))

    # avg секунд → часы
    avg_seconds = func.avg(
        func.extract("epoch", Ticket.closed_at - Ticket.created_at)
    ).label("avg_seconds")

    query = (
        select(Priority.level.label("priority"), avg_seconds)
        .join(Priority, Ticket.priority_id == Priority.id)
        .where(and_(*conditions))
        .group_by(Priority.level)
        .order_by(Priority.level)
    )
    result = await db.execute(query)
    return [
        AvgResolutionItem(
            priority=row.priority.value,
            avg_hours=round(row.avg_seconds / 3600, 2) if row.avg_seconds else None,
        )
        for row in result.all()
    ]


@router.get("/sla-compliance", response_model=SLAComplianceResponse)
async def sla_compliance(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    department_id: int | None = Query(default=None),
    type_id: int | None = Query(default=None),
    priority_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(*_ALLOWED_ROLES)),
):
    conditions = _base_filters(date_from, date_to, department_id, type_id, priority_id)

    query = select(
        func.count(Ticket.id).label("total"),
        func.sum(case((Ticket.sla_violated.is_(False), 1), else_=0)).label("compliant"),
    ).where(and_(*conditions) if conditions else text("1=1"))

    result = await db.execute(query)
    row = result.one()
    total = row.total or 0
    compliant = int(row.compliant or 0)
    rate = round(compliant / total * 100, 1) if total else 0.0
    return SLAComplianceResponse(total=total, compliant=compliant, compliance_rate=rate)


@router.get("/export")
async def export_tickets(
    format: Literal["csv", "xlsx"] = Query(default="csv"),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    department_id: int | None = Query(default=None),
    type_id: int | None = Query(default=None),
    priority_id: int | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(*_ALLOWED_ROLES)),
):
    conditions = _base_filters(date_from, date_to, department_id, type_id, priority_id)

    query = (
        select(
            Ticket.number,
            Ticket.title,
            Ticket.status,
            Priority.name.label("priority"),
            TicketType.name.label("ticket_type"),
            Department.name.label("department"),
            func.concat(User.full_name).label("requester"),
            Ticket.sla_deadline,
            Ticket.sla_violated,
            Ticket.created_at,
            Ticket.closed_at,
        )
        .join(Priority, Ticket.priority_id == Priority.id)
        .join(TicketType, Ticket.type_id == TicketType.id)
        .outerjoin(Department, Ticket.department_id == Department.id)
        .outerjoin(User, Ticket.requester_id == User.id)
        .where(and_(*conditions) if conditions else text("1=1"))
        .order_by(Ticket.created_at.desc())
    )
    result = await db.execute(query)
    rows = result.all()

    headers_row = ["Номер", "Тема", "Статус", "Приоритет", "Тип", "Отдел", "Заявитель",
                   "SLA-дедлайн", "SLA нарушен", "Создана", "Закрыта"]

    def _fmt(v) -> str:
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%d.%m.%Y %H:%M")
        if isinstance(v, bool):
            return "Да" if v else "Нет"
        if hasattr(v, "value"):
            return v.value
        return str(v)

    data = [[_fmt(cell) for cell in row] for row in rows]

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers_row)
        writer.writerows(data)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="tickets.csv"'},
        )

    # xlsx
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1677FF")

    for col_idx, header in enumerate(headers_row, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(data, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Автоширина колонок
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="tickets.xlsx"'},
    )
